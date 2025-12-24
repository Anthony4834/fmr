#!/usr/bin/env python3
"""
Convert FY25 FMR Excel file to CSV (handles corrupted metadata)
"""
import sys
import csv
from openpyxl import load_workbook
from openpyxl.reader.excel import ExcelReader

# Patch the datetime parsing to handle invalid format
original_from_ISO8601 = None

def patched_from_ISO8601(formatted_string):
    """Handle invalid datetime format like '2025- 2-18T20:40:31Z'"""
    try:
        # Try to fix the space issue
        fixed = formatted_string.replace('  ', ' ').replace('- ', '-0').replace('T ', 'T0')
        from openpyxl.utils.datetime import from_ISO8601 as original
        return original(fixed)
    except:
        # If still fails, return a default datetime
        from datetime import datetime
        return datetime.now()

# Monkey patch the datetime parser
import openpyxl.utils.datetime
original_from_ISO8601 = openpyxl.utils.datetime.from_ISO8601
openpyxl.utils.datetime.from_ISO8601 = patched_from_ISO8601

try:
    print("Reading data/FY25_FMRs_revised.xlsx...")
    wb = load_workbook('data/FY25_FMRs_revised.xlsx', read_only=True, data_only=True)
    ws = wb.active
    
    print(f"Sheet: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}")
    
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(cell) if cell is not None else '' for cell in row])
    
    print(f"Writing {len(rows)} rows to data/fmr-2025.csv...")
    with open('data/fmr-2025.csv', 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(rows)
    
    print("✅ Successfully converted to data/fmr-2025.csv")
except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
finally:
    # Restore original function
    if original_from_ISO8601:
        openpyxl.utils.datetime.from_ISO8601 = original_from_ISO8601




