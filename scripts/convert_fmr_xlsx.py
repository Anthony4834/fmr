#!/usr/bin/env python3
"""
Convert an FMR Excel workbook to CSV.

This is a more robust fallback than pandas.read_excel for HUD FMR workbooks that sometimes
contain invalid XML / malformed metadata (which can break pandas/openpyxl in default mode).

Usage:
  python3 scripts/convert_fmr_xlsx.py <input.xlsx> <output.csv>
"""

import sys
import csv

from openpyxl import load_workbook


def patched_from_ISO8601(formatted_string):
    """Handle invalid datetime formats sometimes seen in workbook metadata."""
    try:
        fixed = (
            str(formatted_string)
            .replace("  ", " ")
            .replace("- ", "-0")
            .replace("T ", "T0")
        )
        from openpyxl.utils.datetime import from_ISO8601 as original

        return original(fixed)
    except Exception:
        from datetime import datetime

        return datetime.now()


def main():
    if len(sys.argv) != 3:
        print("Usage: python3 scripts/convert_fmr_xlsx.py <input.xlsx> <output.csv>")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]

    # Monkey patch the datetime parser.
    # NOTE: openpyxl also imports `from_ISO8601` into other modules at import time (e.g. descriptors.base),
    # so we patch both locations to be safe.
    import openpyxl.utils.datetime
    import openpyxl.descriptors.base

    original_from_ISO8601 = openpyxl.utils.datetime.from_ISO8601
    original_base_from_ISO8601 = openpyxl.descriptors.base.from_ISO8601

    openpyxl.utils.datetime.from_ISO8601 = patched_from_ISO8601
    openpyxl.descriptors.base.from_ISO8601 = patched_from_ISO8601

    try:
        print(f"Reading {input_file} with openpyxl...")
        wb = load_workbook(input_file, read_only=True, data_only=True)
        ws = wb.active

        print(f"Sheet: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}")

        with open(output_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                writer.writerow(["" if cell is None else cell for cell in row])
                if idx % 2000 == 0:
                    print(f"  Processed {idx} rows...")

        print(f"✅ Successfully converted to {output_file}")
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback

        traceback.print_exc()
        sys.exit(1)
    finally:
        # Restore original function
        openpyxl.utils.datetime.from_ISO8601 = original_from_ISO8601
        openpyxl.descriptors.base.from_ISO8601 = original_base_from_ISO8601


if __name__ == "__main__":
    main()


